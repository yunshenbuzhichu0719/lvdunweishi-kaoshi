# -*- coding: utf-8 -*-
"""
从「关键岗位人员考试公示题库」目录抽取题库，生成前端可直接 <script src> 引入的 JS 数据包。
科目 A / B / C 全量；科目 D 仅取 2-生态环境监测类、7-卫生计生（产品）（疾控）类。
"""
import os, sys, json, re, hashlib, collections

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

SRC = r'C:\Users\ydyyf\Desktop\关键岗位人员考试公示题库'
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data')

D_DIRS = ['2-生态环境监测类', '7-卫生计生（产品）（疾控）类']

TYPE_MAP = {
    '单选题': 'single', '单选': 'single', '单项选择题': 'single',
    '多选题': 'multi', '多选': 'multi', '多项选择题': 'multi',
    '判断题': 'judge', '判断': 'judge',
}

OPT_LETTERS = ['A', 'B', 'C', 'D', 'E', 'F', 'G']


def norm(v):
    if v is None:
        return ''
    s = str(v)
    s = s.replace('\u00a0', ' ').replace('\ufeff', '')
    s = re.sub(r'[\r\n\t]+', ' ', s)
    return s.strip()


def find_header(ws, scan=15):
    """返回 (header_row_index(0-based), colmap)"""
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=scan, values_only=True)):
        rows.append(row)
    for i, row in enumerate(rows):
        cells = [norm(c) for c in row[:40]]
        joined = '|'.join(cells)
        if '题干' in joined and ('答案' in joined):
            colmap = {}
            for j, c in enumerate(cells):
                if not c:
                    continue
                key = re.sub(r'[（(].*?[)）]', '', c).strip()
                if key.startswith('题干'):
                    colmap['stem'] = j
                elif key.startswith('题型'):
                    colmap['type'] = j
                elif key.startswith('答案'):
                    colmap['answer'] = j
                elif key.startswith('序号'):
                    colmap['no'] = j
                elif key.startswith('选项') and len(key) >= 3:
                    colmap['opt_' + key[2]] = j
                elif '知识点' in key:
                    colmap.setdefault('kp', j)
                elif '解析' in key or '说明' in key:
                    colmap.setdefault('explain', j)
            if 'stem' in colmap and 'answer' in colmap:
                return i, colmap
    return None, None


def parse_sheet(ws, meta):
    hi, cm = find_header(ws)
    if hi is None:
        return []
    out = []
    opt_cols = [(L, cm['opt_' + L]) for L in OPT_LETTERS if 'opt_' + L in cm]
    for i, row in enumerate(ws.iter_rows(min_row=hi + 2, values_only=True)):
        try:
            stem = norm(row[cm['stem']]) if cm['stem'] < len(row) else ''
        except Exception:
            continue
        if not stem:
            continue
        ans = norm(row[cm['answer']]) if cm['answer'] < len(row) else ''
        raw_type = norm(row[cm['type']]) if 'type' in cm and cm['type'] < len(row) else ''
        options = []
        for L, cidx in opt_cols:
            v = norm(row[cidx]) if cidx < len(row) else ''
            if v:
                options.append(v)
            else:
                options.append(None)
        # 去掉尾部空选项
        while options and options[-1] is None:
            options.pop()
        options = ['' if o is None else o for o in options]

        qtype = TYPE_MAP.get(raw_type)
        ansU = re.sub(r'[^A-Za-z对错正确误√×]', '', ans).upper()
        letters = re.findall(r'[A-G]', ansU)
        if qtype is None:
            # 按答案与选项推断
            if len(options) <= 2 and (set(letters) <= {'A', 'B'}):
                qtype = 'judge'
            elif len(letters) > 1:
                qtype = 'multi'
            else:
                qtype = 'single'
        if qtype == 'judge':
            if not letters:
                if any(k in ans for k in ['对', '正确', '√', 'T', 'Y']):
                    letters = ['A']
                elif any(k in ans for k in ['错', '误', '×', 'F', 'N']):
                    letters = ['B']
            options = ['对', '错']
        if not letters:
            continue
        if qtype == 'single' and len(letters) > 1:
            qtype = 'multi'
        if qtype == 'multi' and len(letters) == 1:
            # 题型标注为多选但只有一个答案，仍按多选处理（允许单答案多选）
            pass
        # 答案字母必须落在选项范围内
        maxi = len(options) - 1
        if any(OPT_LETTERS.index(L) > maxi for L in letters):
            continue
        item = {
            'id': '',
            't': {'single': 1, 'multi': 2, 'judge': 3}[qtype],
            'q': stem,
            'o': options,
            'a': ''.join(sorted(set(letters))),
        }
        kp = norm(row[cm['kp']]) if 'kp' in cm and cm['kp'] < len(row) else ''
        if kp:
            item['k'] = kp
        ex = norm(row[cm['explain']]) if 'explain' in cm and cm['explain'] < len(row) else ''
        if ex:
            item['e'] = ex
        item.update(meta)
        out.append(item)
    return out


def load_xlsx(path, meta):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    res = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row is None or ws.max_row < 2:
            continue
        res += parse_sheet(ws, meta)
    wb.close()
    return res


def main():
    banks = []          # 分类元信息
    all_q = []
    stat = []

    def add(items, cat_id):
        n0 = 0
        seen = set()      # 同一题库内去重
        for it in items:
            key = hashlib.md5((cat_id + '#' + it['q'] + '#' + it['a'] + '#' + '|'.join(it['o'])).encode('utf-8')).hexdigest()
            if key in seen:
                continue
            seen.add(key)
            it['id'] = key[:12]
            all_q.append(it)
            n0 += 1
        return n0

    # 科目 A / B / C
    for sub in ['A', 'B', 'C']:
        p = os.path.join(SRC, '科目' + sub, '科目%s.xlsx' % sub)
        items = load_xlsx(p, {'s': sub, 'c': 'S' + sub})
        n = add(items, 'S' + sub)
        banks.append({'id': 'S' + sub, 'subject': sub, 'major': '', 'name': '科目' + sub})
        stat.append(('科目' + sub, len(items), n))

    # 科目 D
    for d in D_DIRS:
        ddir = os.path.join(SRC, '科目D', d)
        major = d
        files = sorted(os.listdir(ddir), key=lambda x: (int(re.match(r'(\d+)', x).group(1)) if re.match(r'(\d+)', x) else 999))
        for f in files:
            if not f.lower().endswith(('.xlsx', '.xls')) or f.startswith('~$'):
                continue
            name = os.path.splitext(f)[0]
            cid = 'SD::%s::%s' % (major, name)
            items = load_xlsx(os.path.join(ddir, f), {'s': 'D', 'c': cid})
            n = add(items, cid)
            banks.append({'id': cid, 'subject': 'D', 'major': major, 'name': name})
            stat.append((major + '/' + name, len(items), n))

    # 统计
    cnt = collections.Counter()
    for q in all_q:
        cnt[(q['c'], q['t'])] += 1
    for b in banks:
        b['n1'] = cnt[(b['id'], 1)]
        b['n2'] = cnt[(b['id'], 2)]
        b['n3'] = cnt[(b['id'], 3)]
        b['total'] = b['n1'] + b['n2'] + b['n3']

    payload = {'version': '2025', 'generated': '', 'banks': banks, 'questions': all_q}
    os.makedirs(OUT, exist_ok=True)
    js = 'window.__KEYPOST_BANK__=' + json.dumps(payload, ensure_ascii=False, separators=(',', ':')) + ';'
    outp = os.path.join(OUT, 'bank-keypost.js')
    with open(outp, 'w', encoding='utf-8') as f:
        f.write(js)

    print('输出:', outp, '%.2f MB' % (os.path.getsize(outp) / 1024 / 1024))
    print('题目总数:', len(all_q))
    for b in banks:
        print('  %-52s 单%-5d 多%-5d 判%-5d 合计%d' % (b['name'][:50], b['n1'], b['n2'], b['n3'], b['total']))


if __name__ == '__main__':
    main()
